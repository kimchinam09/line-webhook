# utils/excel_handler.py
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.utils import get_column_letter

def append_to_excel(file_path, sender_name, department, machine, timestamp, image_path=None):
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Người gửi", "Bộ phận", "Máy", "Thời gian", "Ảnh"])

    row = [sender_name, department, machine, timestamp, image_path or ""]
    ws.append(row)
    wb.save(file_path)


