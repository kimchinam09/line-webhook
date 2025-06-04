# utils/excel_handler.py
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

def write_to_excel(file_path, user_id, department, machine, image_path):
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["User ID", "Department", "Machine", "Date", "Image Path"])
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append([
        user_id,
        department,
        machine,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        image_path
    ])

    wb.save(file_path)
